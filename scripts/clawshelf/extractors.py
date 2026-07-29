from __future__ import annotations

from pathlib import Path
from typing import Protocol
from urllib.error import URLError
from urllib.request import Request, urlopen

import html2text
from openpyxl import load_workbook
import pymupdf
import pymupdf4llm

from .models import ExtractionResult, ProcessingWarning, is_url, source_record, url_source_record


class SourceExtractor(Protocol):
    extraction_method: str

    def supports(self, path: Path) -> bool: ...

    def extract(self, path: Path) -> ExtractionResult: ...


class TextExtractor:
    extensions = {".md", ".markdown", ".txt"}
    extraction_method = "text"

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() in self.extensions

    def extract(self, path: Path) -> ExtractionResult:
        record = source_record(path)
        try:
            content = path.read_text(encoding="utf-8", errors="replace")
        except OSError as exc:
            return ExtractionResult(
                record,
                self.extraction_method,
                "",
                [ProcessingWarning("read_failed", f"{path}: {exc}")],
            )
        return ExtractionResult(record, self.extraction_method, content)


class PdfExtractor:
    extraction_method = "pymupdf4llm-markdown"

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() == ".pdf"

    def extract(self, path: Path) -> ExtractionResult:
        record = source_record(path)
        try:
            with pymupdf.open(path) as document:
                content = self._to_markdown(document, force_ocr=False)
                if not content and document.page_count:
                    content = self._to_markdown(document, force_ocr=True)
        except Exception as exc:
            return ExtractionResult(
                record,
                self.extraction_method,
                "",
                [ProcessingWarning("pdf_read_failed", f"{path}: {exc}")],
            )
        warnings = [] if content else [ProcessingWarning("empty_pdf", "No extractable PDF text.")]
        return ExtractionResult(record, self.extraction_method, content, warnings)

    @staticmethod
    def _to_markdown(document: pymupdf.Document, *, force_ocr: bool) -> str:
        return pymupdf4llm.to_markdown(
            document,
            header=False,
            footer=False,
            ignore_images=False,
            write_images=False,
            use_ocr=True,
            force_ocr=force_ocr,
        ).strip()


class XlsxExtractor:
    extraction_method = "xlsx"

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() == ".xlsx"

    def extract(self, path: Path) -> ExtractionResult:
        record = source_record(path)
        try:
            workbook = load_workbook(path, read_only=True, data_only=False)
        except Exception as exc:
            return ExtractionResult(
                record,
                self.extraction_method,
                "",
                [ProcessingWarning("xlsx_read_failed", f"{path}: {exc}")],
            )
        sections = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            sections.append(f"## Sheet: {sheet.title}")
            if not rows:
                sections.append("(empty sheet)")
                continue
            header = [self._cell(value) for value in rows[0]]
            sections.extend([self._row(header), self._row(["---"] * len(header))])
            sections.extend(self._row([self._cell(value) for value in row]) for row in rows[1:])
        return ExtractionResult(record, self.extraction_method, "\n".join(sections))

    @staticmethod
    def _cell(value: object) -> str:
        return "" if value is None else str(value).replace("|", "\\|").replace("\n", " ")

    @staticmethod
    def _row(values: list[str]) -> str:
        return "| " + " | ".join(values) + " |"


class UrlExtractor:
    timeout_seconds = 20
    user_agent = "ClawShelf/1.0 (+https://github.com/anthropics)"
    extraction_method = "url"

    def supports(self, source: str) -> bool:
        return is_url(source)

    def extract(self, source: str) -> ExtractionResult:
        request = Request(source, headers={"User-Agent": self.user_agent})
        try:
            with urlopen(request, timeout=self.timeout_seconds) as response:
                raw = response.read()
                content_type = response.headers.get_content_type()
                charset = response.headers.get_content_charset() or "utf-8"
        except (URLError, TimeoutError, ValueError, OSError) as exc:
            return ExtractionResult(
                url_source_record(source, b""),
                self.extraction_method,
                "",
                [ProcessingWarning("fetch_failed", f"{source}: {exc}")],
            )

        record = url_source_record(source, raw)
        text = raw.decode(charset, errors="replace")

        if content_type == "text/html":
            converter = html2text.HTML2Text()
            converter.ignore_images = True
            converter.body_width = 0
            content = converter.handle(text).strip()
        else:
            content = text.strip()

        warnings = [] if content else [ProcessingWarning("empty_page", "No extractable text at this URL.")]
        return ExtractionResult(record, self.extraction_method, content, warnings)


class ExtractorRegistry:
    def __init__(
        self,
        extractors: list[SourceExtractor] | None = None,
        url_extractor: UrlExtractor | None = None,
    ):
        self.extractors = extractors or [TextExtractor(), PdfExtractor(), XlsxExtractor()]
        self.url_extractor = url_extractor or UrlExtractor()

    def resolve(self, source: str | Path) -> SourceExtractor | UrlExtractor | None:
        if isinstance(source, str) and is_url(source):
            return self.url_extractor
        path = Path(source)
        for extractor in self.extractors:
            if extractor.supports(path):
                return extractor
        return None

    def expected_extraction_method(self, source: str | Path) -> str | None:
        extractor = self.resolve(source)
        return extractor.extraction_method if extractor is not None else None

    def extract(self, source: str | Path) -> ExtractionResult | None:
        extractor = self.resolve(source)
        return extractor.extract(source) if extractor is not None else None
